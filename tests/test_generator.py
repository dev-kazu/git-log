from __future__ import annotations

from datetime import date
import os
from pathlib import Path
import subprocess
import tempfile
import unittest

from daily_note.generator import (
    Commit,
    DailyNoteError,
    append_daily_sections,
    classify_file,
    first_bullet_after,
    find_git_repos,
    generate_note,
    generate_workspace_note,
    parse_git_log,
    render_daily_markdown,
    update_month_index,
    update_year_workbook,
)


def run(command: list[str], cwd: Path) -> None:
    subprocess.run(command, cwd=cwd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)


def init_repo(repo: Path) -> None:
    run(["git", "init"], repo)
    run(["git", "config", "user.name", "Test User"], repo)
    run(["git", "config", "user.email", "test@example.com"], repo)


def commit_file(repo: Path, path: str, content: str, message: str, day: str, email: str = "test@example.com") -> None:
    file_path = repo / path
    file_path.parent.mkdir(parents=True, exist_ok=True)
    file_path.write_text(content, encoding="utf-8")
    run(["git", "add", path], repo)
    env = os.environ | {
        "GIT_AUTHOR_NAME": "Test User",
        "GIT_AUTHOR_EMAIL": email,
        "GIT_AUTHOR_DATE": f"{day}T10:00:00+09:00",
        "GIT_COMMITTER_NAME": "Test User",
        "GIT_COMMITTER_EMAIL": email,
        "GIT_COMMITTER_DATE": f"{day}T10:00:00+09:00",
    }
    subprocess.run(
        ["git", "commit", "-m", message],
        cwd=repo,
        check=True,
        env=env,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
    )


class GeneratorTest(unittest.TestCase):
    def test_parse_git_log_with_files(self) -> None:
        output = "\x1esha\x1fAlice\x1fa@example.com\x1f2026-05-27T10:00:00+09:00\x1ffeat: add report\nsrc/app.py\nREADME.md\n"

        commits = parse_git_log(output)

        self.assertEqual(
            commits,
            [
                Commit(
                    sha="sha",
                    author_name="Alice",
                    author_email="a@example.com",
                    committed_at="2026-05-27T10:00:00+09:00",
                    subject="feat: add report",
                    files=("src/app.py", "README.md"),
                )
            ],
        )

    def test_classify_file(self) -> None:
        cases = [
            ("src/app.py", "実装"),
            ("tests/test_app.py", "テスト"),
            ("app/foo.spec.js", "テスト"),
            ("README.md", "ドキュメント"),
            (".github/workflows/test.yml", "CI/設定"),
            ("pyproject.toml", "設定"),
        ]
        for path, expected in cases:
            with self.subTest(path=path):
                self.assertEqual(classify_file(path), expected)

    def test_render_daily_markdown_with_empty_commits(self) -> None:
        markdown = render_daily_markdown(date(2026, 5, 27), [], "main")

        self.assertIn("# 日報 2026-05-27", markdown)
        self.assertIn("- 本日のGitコミットはありません。", markdown)
        self.assertIn("## 学んだこと\n- ", markdown)
        self.assertIn("- ブランチ: `main`", markdown)

    def test_append_daily_sections_replaces_empty_placeholders_and_updates_index(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            output_dir = Path(directory) / "daily"
            output_dir.mkdir()
            target = date(2026, 5, 27)
            daily_path = output_dir / "2026-05-27.md"
            daily_path.write_text(render_daily_markdown(target, [], "main"), encoding="utf-8")

            changed = append_daily_sections(
                daily_path,
                {
                    "learned": ["argparse でCLI入力を扱った"],
                    "blocked": ["特になし"],
                    "tomorrow": ["README に使い方を書く"],
                },
            )
            update_month_index(output_dir, target)

            self.assertTrue(changed)
            daily = daily_path.read_text(encoding="utf-8")
            self.assertIn("## 学んだこと\n- argparse でCLI入力を扱った\n", daily)
            self.assertIn("## 詰まったこと・相談したいこと\n- 特になし\n", daily)
            self.assertIn("## 明日やること\n- README に使い方を書く\n", daily)
            self.assertNotIn("## 学んだこと\n- \n", daily)

            index = (output_dir / "2026-05.md").read_text(encoding="utf-8")
            self.assertIn("- 学んだこと: argparse でCLI入力を扱った", index)
            self.assertIn("- 詰まったこと: 特になし", index)

    def test_append_daily_sections_appends_to_existing_items(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            daily_path = Path(directory) / "2026-05-27.md"
            daily_path.write_text(
                "# 日報 2026-05-27\n\n## 学んだこと\n- 既存\n\n## 明日やること\n- \n",
                encoding="utf-8",
            )

            append_daily_sections(daily_path, {"learned": ["追加"], "tomorrow": ["次の作業"]})

            daily = daily_path.read_text(encoding="utf-8")
            self.assertIn("## 学んだこと\n- 既存\n- 追加\n\n## 明日やること", daily)
            self.assertIn("## 明日やること\n- 次の作業\n", daily)

    def test_first_bullet_after_ignores_empty_placeholder(self) -> None:
        content = "## 学んだこと\n- \n\n## 詰まったこと・相談したいこと\n- review flow\n"

        self.assertIsNone(first_bullet_after(content, "## 学んだこと"))
        self.assertEqual(first_bullet_after(content, "## 詰まったこと・相談したいこと"), "review flow")

    def test_generate_note_creates_daily_note_and_month_index(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            repo = tmp_path / "repo"
            repo.mkdir()
            init_repo(repo)
            run(["git", "checkout", "-b", "feature/ABC-123-report"], repo)
            commit_file(repo, "daily_note/app.py", "print('hi')\n", "feat: add daily report", "2026-05-27")
            commit_file(repo, "tests/test_app.py", "def test_app(): pass\n", "test: cover report", "2026-05-27")

            result = generate_note(repo, date(2026, 5, 27), tmp_path / "daily")

            self.assertTrue(result.created)
            daily = result.path.read_text(encoding="utf-8")
            self.assertIn("- 実装: add daily report", daily)
            self.assertIn("- テスト: cover report", daily)
            self.assertIn("実装", daily)
            self.assertIn("テスト", daily)
            self.assertIn("- ブランチ: `feature/ABC-123-report`", daily)
            self.assertIn("- チケット: `ABC-123`", daily)

            index = result.index_path.read_text(encoding="utf-8")
            self.assertIn("# 2026-05 の開発記録", index)
            self.assertIn("- [日報を見る](./2026-05-27.md)", index)
            self.assertIn("- やったこと: 実装: add daily report", index)

    def test_generate_note_creates_year_workbook_with_month_sheets(self) -> None:
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            repo = tmp_path / "repo"
            repo.mkdir()
            init_repo(repo)
            run(["git", "checkout", "-b", "feature/ABC-123-report"], repo)
            commit_file(repo, "daily_note/app.py", "print('hi')\n", "feat: add daily report", "2026-05-27")
            commit_file(repo, "tests/test_app.py", "def test_app(): pass\n", "test: cover report", "2026-05-27")

            result = generate_note(repo, date(2026, 5, 27), tmp_path / "daily")

            self.assertTrue(result.workbook_updated)
            self.assertEqual(result.workbook_path, tmp_path / "daily" / "2026.xlsx")
            workbook = load_workbook(result.workbook_path)
            self.assertEqual(workbook.sheetnames, [f"{month:02d}月" for month in range(1, 13)])
            sheet = workbook["05月"]
            self.assertEqual(sheet.cell(row=2, column=1).value, "2026-05-27")
            self.assertIn("実装: add daily report", sheet.cell(row=2, column=2).value)
            self.assertIn("テスト: cover report", sheet.cell(row=2, column=2).value)
            self.assertIn("実装", sheet.cell(row=2, column=6).value)
            self.assertIn("テスト", sheet.cell(row=2, column=6).value)
            self.assertIn("feature/ABC-123-report", sheet.cell(row=2, column=7).value)
            self.assertEqual(sheet.cell(row=2, column=8).value, (tmp_path / "daily" / "2026-05-27.md").as_posix())

    def test_update_year_workbook_refreshes_existing_day_row(self) -> None:
        from openpyxl import load_workbook

        with tempfile.TemporaryDirectory() as directory:
            output_dir = Path(directory) / "daily"
            output_dir.mkdir()
            target = date(2026, 5, 27)
            daily_path = output_dir / "2026-05-27.md"
            daily_path.write_text(render_daily_markdown(target, [], "main"), encoding="utf-8")

            update_year_workbook(output_dir, target)
            append_daily_sections(daily_path, {"learned": ["初回の学び"], "tomorrow": ["次の作業"]})
            update_year_workbook(output_dir, target)
            append_daily_sections(daily_path, {"learned": ["追加の学び"]})
            update_year_workbook(output_dir, target)

            workbook = load_workbook(output_dir / "2026.xlsx")
            sheet = workbook["05月"]
            self.assertEqual(sheet.max_row, 2)
            self.assertEqual(sheet.cell(row=2, column=1).value, "2026-05-27")
            self.assertEqual(sheet.cell(row=2, column=3).value, "初回の学び\n追加の学び")
            self.assertEqual(sheet.cell(row=2, column=5).value, "次の作業")

    def test_generate_note_filters_to_configured_author(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            repo = tmp_path / "repo"
            repo.mkdir()
            init_repo(repo)
            commit_file(repo, "src/mine.py", "mine\n", "feat: my work", "2026-05-27")
            commit_file(repo, "src/other.py", "other\n", "feat: other work", "2026-05-27", email="other@example.com")

            result = generate_note(repo, date(2026, 5, 27), tmp_path / "daily")

            daily = result.path.read_text(encoding="utf-8")
            self.assertIn("my work", daily)
            self.assertNotIn("other work", daily)

    def test_generate_note_creates_template_without_commits(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            repo = tmp_path / "repo"
            repo.mkdir()
            init_repo(repo)

            result = generate_note(repo, date(2026, 5, 27), tmp_path / "daily")

            daily = result.path.read_text(encoding="utf-8")
            self.assertIn("- 本日のGitコミットはありません。", daily)

    def test_find_git_repos_discovers_multiple_repositories(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            workspace = Path(directory) / "workspace"
            workspace.mkdir()
            repo_one = workspace / "app-one"
            repo_two = workspace / "team" / "app-two"
            repo_one.mkdir()
            repo_two.mkdir(parents=True)
            init_repo(repo_one)
            init_repo(repo_two)

            self.assertEqual(find_git_repos(workspace), [repo_one.resolve(), repo_two.resolve()])

    def test_generate_workspace_note_creates_one_central_note(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            workspace = tmp_path / "workspace"
            workspace.mkdir()
            repo_one = workspace / "app-one"
            repo_two = workspace / "app-two"
            repo_one.mkdir()
            repo_two.mkdir()
            init_repo(repo_one)
            init_repo(repo_two)
            run(["git", "checkout", "-b", "feature/ABC-123-api"], repo_one)
            run(["git", "checkout", "-b", "fix/BUG-9-docs"], repo_two)
            commit_file(repo_one, "src/app.py", "print('one')\n", "feat: add API", "2026-05-27")
            commit_file(repo_two, "README.md", "docs\n", "docs: update setup", "2026-05-27")
            commit_file(repo_two, "src/other.py", "other\n", "feat: other person", "2026-05-27", email="other@example.com")

            result = generate_workspace_note(workspace, date(2026, 5, 27), tmp_path / "daily")

            self.assertTrue(result.created)
            self.assertEqual(result.path, tmp_path / "daily" / "2026-05-27.md")
            self.assertFalse((repo_one / "daily").exists())
            self.assertFalse((repo_two / "daily").exists())
            daily = result.path.read_text(encoding="utf-8")
            self.assertIn("- `app-one`: 実装: add API", daily)
            self.assertIn("- `app-two`: ドキュメント: update setup", daily)
            self.assertNotIn("other person", daily)
            self.assertIn("- `app-one`: 1コミット / `feature/ABC-123-api`", daily)
            self.assertIn("- `app-two`: 1コミット / `fix/BUG-9-docs`", daily)
            self.assertIn("- チケット: `ABC-123`、`BUG-9`", daily)

            index = result.index_path.read_text(encoding="utf-8")
            self.assertIn("- やったこと: `app-one`: 実装: add API", index)

            from openpyxl import load_workbook

            workbook = load_workbook(result.workbook_path)
            sheet = workbook["05月"]
            self.assertIn("`app-one`: 実装: add API", sheet.cell(row=2, column=2).value)
            self.assertIn("`app-two`: ドキュメント: update setup", sheet.cell(row=2, column=2).value)

    def test_generate_workspace_note_errors_without_repositories(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            workspace = Path(directory) / "workspace"
            workspace.mkdir()
            with self.assertRaisesRegex(DailyNoteError, "Gitリポジトリが見つかりません"):
                generate_workspace_note(workspace, date(2026, 5, 27), Path(directory) / "daily")

    def test_generate_note_does_not_overwrite_existing_daily_note(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            repo = tmp_path / "repo"
            repo.mkdir()
            init_repo(repo)
            commit_file(repo, "src/app.py", "hello\n", "feat: add app", "2026-05-27")
            output_dir = tmp_path / "daily"
            output_dir.mkdir()
            daily_path = output_dir / "2026-05-27.md"
            daily_path.write_text("# 日報 2026-05-27\n\n## 学んだこと\n- 手入力\n", encoding="utf-8")

            result = generate_note(repo, date(2026, 5, 27), output_dir)

            self.assertFalse(result.created)
            self.assertEqual(daily_path.read_text(encoding="utf-8"), "# 日報 2026-05-27\n\n## 学んだこと\n- 手入力\n")

    def test_generate_note_requires_git_repo(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            tmp_path = Path(directory)
            with self.assertRaisesRegex(DailyNoteError, "Gitリポジトリではありません"):
                generate_note(tmp_path, date(2026, 5, 27), tmp_path / "daily")


if __name__ == "__main__":
    unittest.main()
